import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Input file paths
CASES_FILE = Path("datasets/cases.csv")
RULE_RESULTS_FILE = Path("datasets/rule_checker_results.csv")
AI_RESULTS_FILE = Path("datasets/ai_diagnosis_results.csv")
REVIEW_FILE = Path("datasets/responsible_ai_log.csv")

# Output file path
OUTPUT_EXCEL = Path("dashboard/NetSega_AI_Dashboard.xlsx")

def create_excel_dashboard():
    print("Generating Cisco NetAcad VIP 2026 AI Track Dashboard Excel File...")
    
    # 1. Load CSV Datasets
    cases_df = pd.read_csv(CASES_FILE)
    rule_df = pd.read_csv(RULE_RESULTS_FILE)
    ai_df = pd.read_csv(AI_RESULTS_FILE)
    resp_df = pd.read_csv(REVIEW_FILE)
    
    # Verify row counts
    total_cases_cnt = len(cases_df)
    assert total_cases_cnt == 31, f"Expected 31 cases, got {total_cases_cnt}"
    assert len(rule_df) == 31, f"Expected 31 rule results, got {len(rule_df)}"
    assert len(ai_df) == 31, f"Expected 31 AI results, got {len(ai_df)}"
    assert len(resp_df) == 31, f"Expected 31 review records, got {len(resp_df)}"
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Color Palette & Styles
    NAVY_HEADER_FILL = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    BLUE_ACCENT_FILL = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
    CARD_BG_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    CARD_HEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Highlight Fills
    SUCCESS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    SUCCESS_FONT = Font(name="Segoe UI", size=10, bold=True, color="166534")
    
    MISMATCH_FILL = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    MISMATCH_FONT = Font(name="Segoe UI", size=10, bold=True, color="9A3412")
    
    EDITED_FILL = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    EDITED_FONT = Font(name="Segoe UI", size=10, bold=True, color="854D0E")
    
    # Fonts
    TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    SUBTITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="E2E8F0")
    SECTION_FONT = Font(name="Segoe UI", size=12, bold=True, color="0B2545")
    HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    SUBHEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="0B2545")
    CARD_TITLE_FONT = Font(name="Segoe UI", size=9, bold=True, color="475569")
    CARD_VALUE_FONT = Font(name="Segoe UI", size=20, bold=True, color="0B2545")
    CARD_SUB_FONT = Font(name="Segoe UI", size=8, italic=True, color="64748B")
    REGULAR_FONT = Font(name="Segoe UI", size=10, color="1E293B")
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    
    # Borders
    THIN_SIDE = Side(border_style="thin", color="CBD5E1")
    THICK_SIDE = Side(border_style="medium", color="0B2545")
    CARD_BORDER_SIDE = Side(border_style="thin", color="94A3B8")
    
    BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    BORDER_CARD = Border(left=CARD_BORDER_SIDE, right=CARD_BORDER_SIDE, top=CARD_BORDER_SIDE, bottom=CARD_BORDER_SIDE)
    BORDER_HEADER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THICK_SIDE, bottom=THICK_SIDE)
    BORDER_BOTTOM_DOUBLE = Border(bottom=Side(border_style="double", color="0B2545"))

    # Alignments
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

    # ==========================================
    # SHEET 1: Summary Dashboard
    # ==========================================
    ws_sum = wb.create_sheet(title="Summary Dashboard")
    ws_sum.sheet_properties.tabColor = "0B2545"
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_sum.merge_cells("A1:L1")
    ws_sum["A1"] = "NETSEGA-AI — AI TRACK DASHBOARD"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A1"].fill = NAVY_HEADER_FILL
    ws_sum["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws_sum.merge_cells("A2:L2")
    ws_sum["A2"] = "Cisco NetAcad VIP 2026 Submission | Automated Network Diagnosis & Responsible AI Evaluation"
    ws_sum["A2"].font = SUBTITLE_FONT
    ws_sum["A2"].fill = NAVY_HEADER_FILL
    ws_sum["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[2].height = 20
    ws_sum.row_dimensions[3].height = 10
    
    # Section 1 Header
    ws_sum["A4"] = "SECTION 1: KEY PERFORMANCE METRICS"
    ws_sum["A4"].font = SECTION_FONT
    ws_sum.row_dimensions[4].height = 24
    
    # KPI Cards Definition (8 Cards arranged in 2 rows of 4 cards)
    # Card 1: Total Cases (B5:C7)
    # Card 2: Successful AI Diagnoses (D5:E7)
    # Card 3: Failed AI Diagnoses (F5:G7)
    # Card 4: Rule Match Rate (H5:I7)
    # Card 5: Human Reviews (B9:C11)
    # Card 6: Accepted Reviews (D9:E11)
    # Card 7: Edited Reviews (F9:G11)
    # Card 8: Human Correction Rate (H9:I11)
    
    kpi_cards = [
        # Row 1 Cards
        {"title": "TOTAL CASES", "formula": "=COUNTA('AI Diagnosis Results'!A2:A32)", "fmt": "#,##0", "start_col": 2, "start_row": 5, "note": "Total Evaluated Network Cases"},
        {"title": "SUCCESSFUL AI DIAGNOSES", "formula": '=COUNTIF(\'AI Diagnosis Results\'!D2:D32, "SUCCESS")', "fmt": "#,##0", "start_col": 4, "start_row": 5, "note": "Gemini API Diagnoses Completed"},
        {"title": "FAILED AI DIAGNOSES", "formula": '=COUNTIF(\'AI Diagnosis Results\'!D2:D32, "<>SUCCESS")', "fmt": "#,##0", "start_col": 6, "start_row": 5, "note": "API Failures or Timeouts"},
        {"title": "RULE MATCH RATE", "formula": "=COUNTIF('Rule Checker Results'!G2:G32, \"MATCH\")/COUNTA('Rule Checker Results'!A2:A32)", "fmt": "0.0%", "start_col": 8, "start_row": 5, "note": "Deterministic Rule Agreement"},
        # Row 2 Cards
        {"title": "HUMAN REVIEWS", "formula": "=COUNTA('Responsible AI - Human Review'!A2:A32)", "fmt": "#,##0", "start_col": 2, "start_row": 9, "note": "Total Cases Reviewed by Expert"},
        {"title": "ACCEPTED REVIEWS", "formula": '=COUNTIF(\'Responsible AI - Human Review\'!E2:E32, "Accepted")', "fmt": "#,##0", "start_col": 4, "start_row": 9, "note": "AI Diagnosis Accepted Intact"},
        {"title": "EDITED REVIEWS", "formula": '=COUNTIF(\'Responsible AI - Human Review\'!E2:E32, "Edited")', "fmt": "#,##0", "start_col": 6, "start_row": 9, "note": "Human Expert Corrections Made"},
        {"title": "HUMAN CORRECTION RATE", "formula": '=COUNTIF(\'Responsible AI - Human Review\'!E2:E32, "Edited")/COUNTA(\'Responsible AI - Human Review\'!A2:A32)', "fmt": "0.0%", "start_col": 8, "start_row": 9, "note": "Percentage of AI Edits Required"},
    ]
    
    for card in kpi_cards:
        r = card["start_row"]
        c = card["start_col"]
        
        # Merge 2 columns for card width
        ws_sum.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
        ws_sum.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
        ws_sum.merge_cells(start_row=r+2, start_column=c, end_row=r+2, end_column=c+1)
        
        c_title = ws_sum.cell(row=r, column=c, value=card["title"])
        c_title.font = CARD_TITLE_FONT
        c_title.fill = CARD_HEADER_FILL
        c_title.alignment = ALIGN_CENTER
        
        c_val = ws_sum.cell(row=r+1, column=c, value=card["formula"])
        c_val.font = CARD_VALUE_FONT
        c_val.fill = CARD_BG_FILL
        c_val.alignment = ALIGN_CENTER
        c_val.number_format = card["fmt"]
        
        c_sub = ws_sum.cell(row=r+2, column=c, value=card["note"])
        c_sub.font = CARD_SUB_FONT
        c_sub.fill = CARD_BG_FILL
        c_sub.alignment = ALIGN_CENTER
        
        # Apply borders around card cells
        for row_idx in range(r, r+3):
            for col_idx in range(c, c+2):
                ws_sum.cell(row=row_idx, column=col_idx).border = BORDER_CARD
    
    ws_sum.row_dimensions[5].height = 18
    ws_sum.row_dimensions[6].height = 28
    ws_sum.row_dimensions[7].height = 16
    ws_sum.row_dimensions[8].height = 10
    ws_sum.row_dimensions[9].height = 18
    ws_sum.row_dimensions[10].height = 28
    ws_sum.row_dimensions[11].height = 16
    ws_sum.row_dimensions[12].height = 15
    
    # Section 2: Summary Data Tables
    ws_sum["A13"] = "SECTION 2: EVALUATION SUMMARY DATA TABLES"
    ws_sum["A13"].font = SECTION_FONT
    ws_sum.row_dimensions[13].height = 24
    
    # Table 1: AI Diagnosis Status Summary (Cols B to D, Rows 15 to 17)
    ws_sum["B15"] = "AI Diagnosis Status"
    ws_sum["C15"] = "Count"
    ws_sum["D15"] = "Percentage"
    for col_letter in ["B", "C", "D"]:
        cell = ws_sum[f"{col_letter}15"]
        cell.font = HEADER_FONT
        cell.fill = BLUE_ACCENT_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    
    ws_sum["B16"] = "SUCCESS"
    ws_sum["C16"] = '=COUNTIF(\'AI Diagnosis Results\'!D2:D32, "SUCCESS")'
    ws_sum["D16"] = '=C16/B6'
    
    ws_sum["B17"] = "FAILED"
    ws_sum["C17"] = '=COUNTIF(\'AI Diagnosis Results\'!D2:D32, "<>SUCCESS")'
    ws_sum["D17"] = '=C17/B6'
    
    ws_sum["B18"] = "Total"
    ws_sum["C18"] = '=SUM(C16:C17)'
    ws_sum["D18"] = '=SUM(D16:D17)'
    
    for r in range(16, 19):
        ws_sum[f"B{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"B{r}"].border = BORDER_ALL
        ws_sum[f"C{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"C{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"C{r}"].number_format = "#,##0"
        ws_sum[f"C{r}"].border = BORDER_ALL
        ws_sum[f"D{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"D{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"D{r}"].number_format = "0.0%"
        ws_sum[f"D{r}"].border = BORDER_ALL
    
    # Table 2: Human Review Decisions (Cols F to H, Rows 15 to 17)
    ws_sum["F15"] = "Human Review Decision"
    ws_sum["G15"] = "Count"
    ws_sum["H15"] = "Percentage"
    for col_letter in ["F", "G", "H"]:
        cell = ws_sum[f"{col_letter}15"]
        cell.font = HEADER_FONT
        cell.fill = BLUE_ACCENT_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        
    ws_sum["F16"] = "Accepted"
    ws_sum["G16"] = '=COUNTIF(\'Responsible AI - Human Review\'!E2:E32, "Accepted")'
    ws_sum["H16"] = '=G16/B6'
    
    ws_sum["F17"] = "Edited"
    ws_sum["G17"] = '=COUNTIF(\'Responsible AI - Human Review\'!E2:E32, "Edited")'
    ws_sum["H17"] = '=G17/B6'
    
    ws_sum["F18"] = "Total"
    ws_sum["G18"] = '=SUM(G16:G17)'
    ws_sum["H18"] = '=SUM(H16:H17)'
    
    for r in range(16, 19):
        ws_sum[f"F{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"F{r}"].border = BORDER_ALL
        ws_sum[f"G{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"G{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"G{r}"].number_format = "#,##0"
        ws_sum[f"G{r}"].border = BORDER_ALL
        ws_sum[f"H{r}"].font = BOLD_FONT if r == 18 else REGULAR_FONT
        ws_sum[f"H{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"H{r}"].number_format = "0.0%"
        ws_sum[f"H{r}"].border = BORDER_ALL

    # Table 3: Rule Checker Comparison (Cols J to L, Rows 15 to 18)
    ws_sum["J15"] = "Rule Checker Comparison"
    ws_sum["K15"] = "Count"
    ws_sum["L15"] = "Percentage"
    for col_letter in ["J", "K", "L"]:
        cell = ws_sum[f"{col_letter}15"]
        cell.font = HEADER_FONT
        cell.fill = BLUE_ACCENT_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        
    ws_sum["J16"] = "MATCH"
    ws_sum["K16"] = '=COUNTIF(\'Rule Checker Results\'!G2:G32, "MATCH")'
    ws_sum["L16"] = '=K16/B6'
    
    ws_sum["J17"] = "MISMATCH"
    ws_sum["K17"] = '=COUNTIF(\'Rule Checker Results\'!G2:G32, "MISMATCH")'
    ws_sum["L17"] = '=K17/B6'
    
    ws_sum["J18"] = "NO_DETECTION"
    ws_sum["K18"] = '=COUNTIF(\'Rule Checker Results\'!G2:G32, "NO_DETECTION")'
    ws_sum["L18"] = '=K18/B6'
    
    ws_sum["J19"] = "Total"
    ws_sum["K19"] = '=SUM(K16:K18)'
    ws_sum["L19"] = '=SUM(L16:L18)'
    
    for r in range(16, 20):
        ws_sum[f"J{r}"].font = BOLD_FONT if r == 19 else REGULAR_FONT
        ws_sum[f"J{r}"].border = BORDER_ALL
        ws_sum[f"K{r}"].font = BOLD_FONT if r == 19 else REGULAR_FONT
        ws_sum[f"K{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"K{r}"].number_format = "#,##0"
        ws_sum[f"K{r}"].border = BORDER_ALL
        ws_sum[f"L{r}"].font = BOLD_FONT if r == 19 else REGULAR_FONT
        ws_sum[f"L{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"L{r}"].number_format = "0.0%"
        ws_sum[f"L{r}"].border = BORDER_ALL

    # Table 4: Confidence Distribution Table (Cols N to P, Rows 15 to 19)
    ws_sum["N15"] = "Confidence Bin"
    ws_sum["O15"] = "Count"
    ws_sum["P15"] = "Percentage"
    for col_letter in ["N", "O", "P"]:
        cell = ws_sum[f"{col_letter}15"]
        cell.font = HEADER_FONT
        cell.fill = BLUE_ACCENT_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        
    ws_sum["N16"] = "< 0.80"
    ws_sum["O16"] = '=COUNTIF(\'AI Diagnosis Results\'!C2:C32, "<0.80")'
    ws_sum["P16"] = '=O16/B6'
    
    ws_sum["N17"] = "0.80 - 0.89"
    ws_sum["O17"] = '=COUNTIFS(\'AI Diagnosis Results\'!C2:C32, ">=0.80", \'AI Diagnosis Results\'!C2:C32, "<0.90")'
    ws_sum["P17"] = '=O17/B6'
    
    ws_sum["N18"] = "0.90 - 0.95"
    ws_sum["O18"] = '=COUNTIFS(\'AI Diagnosis Results\'!C2:C32, ">=0.90", \'AI Diagnosis Results\'!C2:C32, "<=0.95")'
    ws_sum["P18"] = '=O18/B6'
    
    ws_sum["N19"] = "0.96 - 1.00"
    ws_sum["O19"] = '=COUNTIFS(\'AI Diagnosis Results\'!C2:C32, ">0.95", \'AI Diagnosis Results\'!C2:C32, "<=1.00")'
    ws_sum["P19"] = '=O19/B6'
    
    ws_sum["N20"] = "Total"
    ws_sum["O20"] = '=SUM(O16:O19)'
    ws_sum["P20"] = '=SUM(P16:P19)'
    
    for r in range(16, 21):
        ws_sum[f"N{r}"].font = BOLD_FONT if r == 20 else REGULAR_FONT
        ws_sum[f"N{r}"].border = BORDER_ALL
        ws_sum[f"O{r}"].font = BOLD_FONT if r == 20 else REGULAR_FONT
        ws_sum[f"O{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"O{r}"].number_format = "#,##0"
        ws_sum[f"O{r}"].border = BORDER_ALL
        ws_sum[f"P{r}"].font = BOLD_FONT if r == 20 else REGULAR_FONT
        ws_sum[f"P{r}"].alignment = ALIGN_RIGHT
        ws_sum[f"P{r}"].number_format = "0.0%"
        ws_sum[f"P{r}"].border = BORDER_ALL

    # Section 3 Header: Visual Analytics
    ws_sum["A22"] = "SECTION 3: VISUAL ANALYTICS & EVALUATION CHARTS"
    ws_sum["A22"].font = SECTION_FONT
    ws_sum.row_dimensions[22].height = 24
    
    # -------------------------------------------------------------
    # CHART 1: AI Success vs Failure (PieChart)
    # -------------------------------------------------------------
    pie1 = PieChart()
    pie1.title = "AI Diagnosis Status (Success vs Failure)"
    pie1.style = 10
    pie1.height = 7.5
    pie1.width = 11
    
    labels1 = Reference(ws_sum, min_col=2, min_row=16, max_row=17)
    data1 = Reference(ws_sum, min_col=3, min_row=15, max_row=17)
    pie1.add_data(data1, titles_from_data=True)
    pie1.set_categories(labels1)
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showVal = True
    ws_sum.add_chart(pie1, "B24")

    # -------------------------------------------------------------
    # CHART 2: Human Accepted vs Edited (PieChart)
    # -------------------------------------------------------------
    pie2 = PieChart()
    pie2.title = "Human Review Decisions (Accepted vs Edited)"
    pie2.style = 10
    pie2.height = 7.5
    pie2.width = 11
    
    labels2 = Reference(ws_sum, min_col=6, min_row=16, max_row=17)
    data2 = Reference(ws_sum, min_col=7, min_row=15, max_row=17)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showVal = True
    ws_sum.add_chart(pie2, "H24")

    # -------------------------------------------------------------
    # CHART 3: Rule Match Rate Breakdown (BarChart)
    # -------------------------------------------------------------
    bar3 = BarChart()
    bar3.type = "col"
    bar3.style = 10
    bar3.title = "Rule Checker Performance Breakdown"
    bar3.y_axis.title = "Number of Cases"
    bar3.x_axis.title = "Comparison Category"
    bar3.height = 7.5
    bar3.width = 11
    bar3.legend = None
    
    labels3 = Reference(ws_sum, min_col=10, min_row=16, max_row=18)
    data3 = Reference(ws_sum, min_col=11, min_row=15, max_row=18)
    bar3.add_data(data3, titles_from_data=True)
    bar3.set_categories(labels3)
    bar3.dataLabels = DataLabelList()
    bar3.dataLabels.showVal = True
    ws_sum.add_chart(bar3, "B39")

    # -------------------------------------------------------------
    # CHART 4: Confidence Distribution (BarChart)
    # -------------------------------------------------------------
    bar4 = BarChart()
    bar4.type = "col"
    bar4.style = 10
    bar4.title = "AI Diagnosis Confidence Distribution"
    bar4.y_axis.title = "Number of Cases"
    bar4.x_axis.title = "Confidence Bins"
    bar4.height = 7.5
    bar4.width = 11
    bar4.legend = None
    
    labels4 = Reference(ws_sum, min_col=14, min_row=16, max_row=19)
    data4 = Reference(ws_sum, min_col=15, min_row=15, max_row=19)
    bar4.add_data(data4, titles_from_data=True)
    bar4.set_categories(labels4)
    bar4.dataLabels = DataLabelList()
    bar4.dataLabels.showVal = True
    ws_sum.add_chart(bar4, "H39")


    # ==========================================
    # SHEET 2: AI Diagnosis Results
    # ==========================================
    ws_ai = wb.create_sheet(title="AI Diagnosis Results")
    ws_ai.sheet_properties.tabColor = "0070D2"
    ws_ai.views.sheetView[0].showGridLines = True
    
    ai_headers = [
        "Case ID", "AI Diagnosis", "Confidence", "Status", 
        "OSI Layer", "Concept", "Next Command", "Human Review Required"
    ]
    
    ws_ai.append(ai_headers)
    ws_ai.row_dimensions[1].height = 26
    for col_num, header in enumerate(ai_headers, 1):
        cell = ws_ai.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    for row_idx, row in ai_df.iterrows():
        r = row_idx + 2
        ws_ai.row_dimensions[r].height = 20
        fill = ZEBRA_FILL if row_idx % 2 == 1 else WHITE_FILL
        
        ws_ai.cell(row=r, column=1, value=str(row["case_id"])).alignment = ALIGN_CENTER
        ws_ai.cell(row=r, column=2, value=str(row["root_cause"])).alignment = ALIGN_LEFT
        
        c_conf = ws_ai.cell(row=r, column=3, value=float(row["confidence"]))
        c_conf.alignment = ALIGN_RIGHT
        c_conf.number_format = "0.00"
        
        c_stat = ws_ai.cell(row=r, column=4, value=str(row["status"]))
        c_stat.alignment = ALIGN_CENTER
        if row["status"] == "SUCCESS":
            c_stat.fill = SUCCESS_FILL
            c_stat.font = SUCCESS_FONT
            
        ws_ai.cell(row=r, column=5, value=str(row["osi_layer"])).alignment = ALIGN_CENTER
        ws_ai.cell(row=r, column=6, value=str(row["concept"])).alignment = ALIGN_LEFT
        ws_ai.cell(row=r, column=7, value=str(row["next_command"])).alignment = ALIGN_LEFT
        
        c_rev = ws_ai.cell(row=r, column=8, value=str(row["human_review_required"]))
        c_rev.alignment = ALIGN_CENTER

        for col_num in range(1, 9):
            c_item = ws_ai.cell(row=r, column=col_num)
            if col_num != 4 or row["status"] != "SUCCESS":
                c_item.fill = fill
                c_item.font = REGULAR_FONT
            c_item.border = BORDER_ALL


    # ==========================================
    # SHEET 3: Rule Checker Results
    # ==========================================
    ws_rule = wb.create_sheet(title="Rule Checker Results")
    ws_rule.sheet_properties.tabColor = "134074"
    ws_rule.views.sheetView[0].showGridLines = True
    
    rule_headers = [
        "Case ID", "Expected Fault", "Concept Tag", "Detected Concepts",
        "Rule Findings", "Issues Found", "Comparison", "Status"
    ]
    
    ws_rule.append(rule_headers)
    ws_rule.row_dimensions[1].height = 26
    for col_num, header in enumerate(rule_headers, 1):
        cell = ws_rule.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    for row_idx, row in rule_df.iterrows():
        r = row_idx + 2
        ws_rule.row_dimensions[r].height = 20
        fill = ZEBRA_FILL if row_idx % 2 == 1 else WHITE_FILL
        
        ws_rule.cell(row=r, column=1, value=str(row["case_id"])).alignment = ALIGN_CENTER
        ws_rule.cell(row=r, column=2, value=str(row["expected_fault"])).alignment = ALIGN_LEFT
        ws_rule.cell(row=r, column=3, value=str(row["concept_tag"])).alignment = ALIGN_CENTER
        ws_rule.cell(row=r, column=4, value=str(row["detected_concepts"])).alignment = ALIGN_LEFT
        ws_rule.cell(row=r, column=5, value=str(row["rule_findings"])).alignment = ALIGN_LEFT
        
        c_iss = ws_rule.cell(row=r, column=6, value=int(row["issues_found"]))
        c_iss.alignment = ALIGN_RIGHT
        c_iss.number_format = "#,##0"
        
        c_comp = ws_rule.cell(row=r, column=7, value=str(row["comparison"]))
        c_comp.alignment = ALIGN_CENTER
        if row["comparison"] == "MATCH":
            c_comp.fill = SUCCESS_FILL
            c_comp.font = SUCCESS_FONT
        elif row["comparison"] == "MISMATCH":
            c_comp.fill = MISMATCH_FILL
            c_comp.font = MISMATCH_FONT
            
        c_stat = ws_rule.cell(row=r, column=8, value=str(row["status"]))
        c_stat.alignment = ALIGN_CENTER

        for col_num in range(1, 9):
            c_item = ws_rule.cell(row=r, column=col_num)
            if col_num != 7 or row["comparison"] not in ["MATCH", "MISMATCH"]:
                c_item.fill = fill
                c_item.font = REGULAR_FONT
            c_item.border = BORDER_ALL


    # ==========================================
    # SHEET 4: Responsible AI - Human Review
    # ==========================================
    ws_resp = wb.create_sheet(title="Responsible AI - Human Review")
    ws_resp.sheet_properties.tabColor = "0D9488"
    ws_resp.views.sheetView[0].showGridLines = True
    
    resp_headers = [
        "Case ID", "AI Diagnosis", "AI Confidence", "Expected Fault",
        "Human Decision", "Corrected Diagnosis", "Reviewer Note", "Review Timestamp"
    ]
    
    ws_resp.append(resp_headers)
    ws_resp.row_dimensions[1].height = 26
    for col_num, header in enumerate(resp_headers, 1):
        cell = ws_resp.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    for row_idx, row in resp_df.iterrows():
        r = row_idx + 2
        ws_resp.row_dimensions[r].height = 20
        fill = ZEBRA_FILL if row_idx % 2 == 1 else WHITE_FILL
        
        ws_resp.cell(row=r, column=1, value=str(row["case_id"])).alignment = ALIGN_CENTER
        ws_resp.cell(row=r, column=2, value=str(row["ai_diagnosis"])).alignment = ALIGN_LEFT
        
        c_conf = ws_resp.cell(row=r, column=3, value=float(row["ai_confidence"]))
        c_conf.alignment = ALIGN_RIGHT
        c_conf.number_format = "0.00"
        
        ws_resp.cell(row=r, column=4, value=str(row["expected_fault"])).alignment = ALIGN_LEFT
        
        c_dec = ws_resp.cell(row=r, column=5, value=str(row["human_decision"]))
        c_dec.alignment = ALIGN_CENTER
        if row["human_decision"] == "Accepted":
            c_dec.fill = SUCCESS_FILL
            c_dec.font = SUCCESS_FONT
        elif row["human_decision"] == "Edited":
            c_dec.fill = EDITED_FILL
            c_dec.font = EDITED_FONT
            
        ws_resp.cell(row=r, column=6, value=str(row["corrected_diagnosis"])).alignment = ALIGN_LEFT
        ws_resp.cell(row=r, column=7, value=str(row["reviewer_note"])).alignment = ALIGN_LEFT
        ws_resp.cell(row=r, column=8, value=str(row["review_timestamp"])).alignment = ALIGN_CENTER

        for col_num in range(1, 9):
            c_item = ws_resp.cell(row=r, column=col_num)
            if col_num != 5 or row["human_decision"] not in ["Accepted", "Edited"]:
                c_item.fill = fill
                c_item.font = REGULAR_FONT
            c_item.border = BORDER_ALL

    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Ignore merged cells in Title row of Summary Dashboard
                if ws.title == "Summary Dashboard" and cell.row in [1, 2]:
                    continue
                val_str = str(cell.value or "")
                if val_str.startswith("="):
                    val_str = "100.0%"  # Placeholder length for formula results
                max_len = max(max_len, len(val_str))
            
            # Set minimum width and padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Custom tweaks for specific columns in Summary Dashboard
    if "Summary Dashboard" in wb.sheetnames:
        ws_s = wb["Summary Dashboard"]
        ws_s.column_dimensions["A"].width = 5
        ws_s.column_dimensions["B"].width = 16
        ws_s.column_dimensions["C"].width = 16
        ws_s.column_dimensions["D"].width = 16
        ws_s.column_dimensions["E"].width = 5
        ws_s.column_dimensions["F"].width = 22
        ws_s.column_dimensions["G"].width = 14
        ws_s.column_dimensions["H"].width = 16
        ws_s.column_dimensions["I"].width = 5
        ws_s.column_dimensions["J"].width = 24
        ws_s.column_dimensions["K"].width = 14
        ws_s.column_dimensions["L"].width = 16
        ws_s.column_dimensions["M"].width = 5
        ws_s.column_dimensions["N"].width = 18
        ws_s.column_dimensions["O"].width = 14
        ws_s.column_dimensions["P"].width = 16

    # Save Workbook
    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_EXCEL)
    print(f"Successfully generated Excel Dashboard at: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    create_excel_dashboard()
